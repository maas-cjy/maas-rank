#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MaaS Rank —— 榜单数据自动更新脚本
================================================
数据源（均为公开接口，无需密钥）：

  1. LMArena Elo（全球人类盲测 Elo）
     HuggingFace 数据集 lmarena-ai/leaderboard-dataset
     https://datasets-server.huggingface.co/rows?dataset=lmarena-ai/leaderboard-dataset&config=text&split=latest

  2. SuperCLUE 智能指数（中文能力）
     https://www.superclueai.com/data/generalboard/<月份>.xlsx   （例：2026年7月.xlsx）

  3. SuperCLUE 输入 / 输出价格（人民币 · 元/百万 tokens）
     https://www.superclueai.com/data/latency_and_price/<月份>_2.xlsx

用法：
  python scripts/fetch_data.py                 # 默认更新 data/models.json
  python scripts/fetch_data.py --dry-run       # 只打印变更，不写文件

行为设计：
  - 只更新「成功抓取且能匹配到模型」的字段；抓取失败 / 未匹配时保留旧值，
    因此任何数据源挂掉都不会破坏网站。
  - 字段被权威来源更新后，自动移除该字段的 est（估算）标记。
  - 无论抓取成功与否，退出码都为 0（文件读写失败除外），便于 CI 使用。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = ROOT / "data" / "models.json"
DIFF_FILE = ROOT / "data" / ".last_update.json"
HISTORY_FILE = ROOT / "data" / "history.json"
TIMEOUT = 25

# 历史快照仅归档以下字段（自动更新只涉及这几个指标）
HISTORY_FIELDS = ("id", "name", "elo", "superclue", "bench", "priceIn", "priceOut")

# SuperCLUE 总排行榜中的六维子分列名（用于计算「综合能力分」= 可用子分均值）
DIM_KEYS = [
    ("math", "数学推理"),
    ("hallu", "幻觉控制"),
    ("science", "科学推理"),
    ("ifollow", "精确指令遵循"),
    ("coding", "智能体编程"),
    ("plan", "智能体任务规划"),
]

# ---------------------------------------------------------------------------
# 模型别名表：models.json 的 id -> 各数据源中的名称特征（子串匹配，忽略大小写）
# ---------------------------------------------------------------------------
MODEL_ALIASES = {
    "claude-opus-5":      ["claude opus 5", "claude-opus-5", "claude-opus-4-7"],
    "claude-opus-4.8":    ["claude opus 4.8", "claude-opus-4.8"],
    "claude-sonnet-4.6":  ["claude sonnet 4.6", "claude-sonnet-4.6"],
    "gpt-5.5":            ["gpt-5.5", "gpt 5.5"],
    "gpt-5.4":            ["gpt-5.4", "gpt 5.4"],
    "gemini-3.1-pro":     ["gemini-3.1-pro", "gemini 3.1 pro"],
    "gemini-3.5-flash":   ["gemini-3.5-flash", "gemini 3.5 flash"],
    "grok-4.5":           ["grok 4.5", "grok-4.5", "grok-4"],
    "llama-4-maverick":   ["llama 4 maverick", "llama-4-maverick"],
    "qwen3.8-max":        ["qwen3.8-max", "qwen3.8 max"],
    "qwen3.7-max":        ["qwen3.7-max", "qwen3.7 max"],
    "deepseek-v4-pro":    ["deepseek-v4-pro", "deepseek v4 pro"],
    "deepseek-v4-flash":  ["deepseek-v4-flash", "deepseek v4 flash"],
    "kimi-k3":            ["kimi-k3", "kimi k3", "kimi-k2.6"],
    "glm-5.2":            ["glm-5.2", "glm 5.2", "glm-5.1"],
    "doubao-seed-2.1-pro": ["doubao-seed-2.1-pro", "doubao seed 2.1", "doubao-seed-2.0-pro"],
    "ernie-5.1":          ["ernie 5.1", "ernie-5.1", "文心"],
    "mimo-v2.5-pro":      ["mimo-v2.5-pro", "mimo v2.5"],
    "minimax-m3":         ["minimax-m3", "minimax m3"],
    "hunyuan-hy3":        ["hy3", "混元"],
}

# SuperCLUE 已知的发布月份（前端写死的可选项），新月份上线后会更新 JS，因此脚本还会
# 动态往前推月份作为候选，取「最新可用」的一份。
SC_KNOWN_DATES = [
    "2026年7月", "2026年5月", "2026年3月", "2025年度测评", "2025年11月",
    "2025年9月", "2025年7月", "2025年5月", "2025年3月", "2024年12月",
    "2024年10月", "2024年8月", "2024年6月", "2024年4月", "2024年2月",
    "2023年12月", "2023年11月", "2023年10月", "2023年9月",
]

SC_BASE = "https://www.superclueai.com/data"
HF_DS = "https://datasets-server.huggingface.co/rows"


def hf_rows_url(dataset: str, config: str, split: str, offset: int, length: int) -> str:
    import urllib.parse as up

    return (
        f"{HF_DS}?dataset={up.quote(dataset)}&config={up.quote(config)}"
        f"&split={up.quote(split)}&offset={offset}&length={length}"
    )


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def http_get(url: str, retries: int = 2) -> bytes:
    """带重试的 GET，返回原始字节；彻底失败抛异常。"""
    last_err: Exception | None = None
    for i in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (MaaS-Rank updater)"})
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                return resp.read()
        except Exception as e:  # noqa: BLE001
            last_err = e
            log(f"  ⚠ 第 {i + 1} 次请求失败: {e}")
    raise RuntimeError(f"GET {url} 失败: {last_err}")


def to_num(v) -> float | None:
    """把 Excel / JSON 里的值转成 float；'-'、None、空串返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("¥", "").replace("$", "")
    if s in ("", "-", "--", "None", "null"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None


# ---------------------------------------------------------------------------
# 1. LMArena Elo —— HuggingFace datasets-server
# ---------------------------------------------------------------------------
def fetch_lmarena() -> tuple[dict[str, float], list[dict]]:
    """返回 (elo_map: {模型名小写: elo}, rows: 全量原始行)。失败返回 ({}, [])。"""
    log("→ 抓取 LMArena Elo (HuggingFace lmarena-ai/leaderboard-dataset)")
    dataset = "lmarena-ai/leaderboard-dataset"
    out: dict[str, float] = {}
    rows: list[dict] = []
    try:
        offset, page_size = 0, 100
        while True:
            url = hf_rows_url(dataset, "text", "latest", offset, page_size)
            raw = http_get(url)
            data = json.loads(raw)
            page = data.get("rows", [])
            if not page:
                break
            for row in page:
                f = row.get("row", {})
                category = str(f.get("category", "")).lower()
                if category and category != "overall":
                    continue
                name = f.get("model_name") or f.get("model") or f.get("name")
                rating = to_num(f.get("rating") if f.get("rating") is not None else f.get("score"))
                if name and rating is not None:
                    key = str(name).lower().strip()
                    out[key] = rating
                    rows.append({
                        "name": key,
                        "org": str(f.get("organization") or "").strip(),
                        "license": str(f.get("license") or "").strip(),
                        "rating": rating,
                    })
            offset += len(page)
            if len(page) < page_size:
                break
    except Exception as e:  # noqa: BLE001
        log(f"  ✗ LMArena 抓取失败：{e}（保留旧 Elo）")
        return {}, []
    log(f"  ✓ 拿到 {len(out)} 个模型的 Elo（overall 分类）")
    return out, rows


# ---------------------------------------------------------------------------
# 2/3. SuperCLUE —— 总榜 xlsx + 价格 xlsx
# ---------------------------------------------------------------------------
def sc_candidate_dates(limit: int = 14) -> list[str]:
    """动态候选（从当前月往前推）+ 已知列表，去重保序。"""
    today = date.today()
    dyn = [f"{today.year}年{today.month}月"]
    y, m = today.year, today.month
    for _ in range(limit):
        m -= 1
        if m == 0:
            y, m = y - 1, 12
        dyn.append(f"{y}年{m}月")
    seen, out = set(), []
    for d in dyn + SC_KNOWN_DATES:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out


def sc_load_xlsx(url: str):
    """下载 xlsx 并用 openpyxl 解析；失败返回 None。"""
    import io

    import openpyxl

    try:
        raw = http_get(url)
        if raw[:2] != b"PK":
            log(f"  ⚠ {url} 不是有效 xlsx（可能是 404 页面），跳过")
            return None
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        return wb
    except Exception as e:  # noqa: BLE001
        log(f"  ⚠ 解析失败 {url}: {e}")
        return None


def sheet_to_rows(ws) -> list[list]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell for cell in row])
    return rows


def fetch_superclue() -> tuple[dict[str, dict], dict[str, dict], str]:
    """
    返回 (总榜: {模型名小写: {"total": 总分, "dims": {六维子分}}}, 价格, 数据月份)
    失败时返回空 dict。
    """
    scores: dict[str, dict] = {}
    prices: dict[str, dict] = {}
    used_date = ""
    for d in sc_candidate_dates():
        gb_url = f"{SC_BASE}/generalboard/{urllib.parse.quote(d)}.xlsx"
        wb = sc_load_xlsx(gb_url)
        if wb is None:
            continue
        # —— 总榜：取「总排行榜」sheet 的总分与六维子分（按表头定位，列序变化不破坏） ——
        try:
            sheet = wb["总排行榜"]
        except KeyError:
            wb.close()
            continue
        rows = sheet_to_rows(sheet)
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        def col(keyword: str) -> int:
            for i, h in enumerate(headers):
                if keyword.lower() in h:
                    return i
            return -1

        i_name, i_total = col("模型名称"), col("总分")
        i_dims = {k: col(kw) for k, kw in DIM_KEYS}
        for r in rows[1:]:
            if i_name < 0 or len(r) <= i_name or not r[i_name]:
                continue
            key = str(r[i_name]).strip().lower()
            entry: dict = {"dims": {}}
            if i_total >= 0 and len(r) > i_total:
                t = to_num(r[i_total])
                if t is not None:
                    entry["total"] = t
            for k, i in i_dims.items():
                if i >= 0 and len(r) > i:
                    v = to_num(r[i])
                    if v is not None:
                        entry["dims"][k] = v
            if "total" in entry or entry["dims"]:
                scores[key] = entry
        wb.close()

        # —— 价格：latency_and_price 用 `<月份>_2.xlsx`（前端约定），退化为 _1 / 裸名 ——
        for suffix in ("_2", "_1", ""):
            lp_url = f"{SC_BASE}/latency_and_price/{urllib.parse.quote(d + suffix)}.xlsx"
            wb2 = sc_load_xlsx(lp_url)
            if wb2 is None:
                continue
            sheet2 = wb2[wb2.sheetnames[0]]
            rows2 = sheet_to_rows(sheet2)
            headers = [str(h).lower() if h else "" for h in rows2[0]]
            def col(keyword: str) -> int:
                for i, h in enumerate(headers):
                    if keyword in h:
                        return i
                return -1
            i_in, i_out = col("input"), col("output")
            for r in rows2[1:]:
                name = r[0] if r else None
                if not name:
                    continue
                key = str(name).strip().lower()
                entry = prices.setdefault(key, {})
                if i_in >= 0 and len(r) > i_in:
                    v = to_num(r[i_in])
                    if v is not None:
                        entry["priceIn"] = v
                if i_out >= 0 and len(r) > i_out:
                    v = to_num(r[i_out])
                    if v is not None:
                        entry["priceOut"] = v
            wb2.close()
            break  # 该月份价格文件命中一次即可
        used_date = d
        log(f"  ✓ SuperCLUE 数据月份：{d}（智能指数 {len(scores)} 个，价格 {len(prices)} 个）")
        break
    else:
        log("  ✗ SuperCLUE 所有候选月份均抓取失败（保留旧值）")
    return scores, prices, used_date


# ---------------------------------------------------------------------------
# 合并
# ---------------------------------------------------------------------------
# LMArena organization -> 厂商显示名；中国厂商集合（用于区域徽章）
ORG_NAMES = {
    "anthropic": "Anthropic", "openai": "OpenAI", "google": "Google",
    "alibaba": "阿里云", "meta": "Meta", "xai": "xAI", "deepseek": "DeepSeek",
    "zai": "智谱", "zhipu": "智谱", "baidu": "百度", "moonshot": "月之暗面",
    "xiaomi": "小米", "bytedance": "字节跳动", "minimax": "MiniMax",
    "mistral": "Mistral AI", "nvidia": "NVIDIA", "tencent": "腾讯",
    "meituan": "美团", "amazon": "Amazon", "thinky": "Thinky",
    "01-ai": "零一万物", "stepfun": "阶跃星辰",
}
CN_ORGS = {"alibaba", "deepseek", "zai", "zhipu", "baidu", "moonshot", "xiaomi",
           "bytedance", "minimax", "tencent", "meituan", "01-ai", "stepfun"}

# 专有名词段 -> 规范大小写（用于生成新模型展示名）
PROPER_NOUNS = {
    "claude": "Claude", "opus": "Opus", "sonnet": "Sonnet", "haiku": "Haiku",
    "gpt": "GPT", "chatgpt": "ChatGPT", "gemini": "Gemini", "gemma": "Gemma",
    "qwen": "Qwen", "deepseek": "DeepSeek", "kimi": "Kimi", "glm": "GLM",
    "grok": "Grok", "llama": "Llama", "mistral": "Mistral", "minimax": "MiniMax",
    "ernie": "ERNIE", "mimo": "MiMo", "doubao": "Doubao", "hunyuan": "混元",
    "hy": "混元", "nova": "Nova", "nemotron": "Nemotron", "muse": "Muse",
    "spark": "Spark", "longcat": "LongCat", "nvidia": "NVIDIA", "moonshot": "Moonshot",
    "xai": "xAI", "openai": "OpenAI", "amazon": "Amazon", "gemma": "Gemma",
}

# LMArena 模型名的「档位/状态/快照」剥离规则（用于家族归并：同一模型的
# -high/-max/-thinking/-preview/日期快照等变体合并为一个条目，取 Elo 最高者）
STRIP_RE = [
    re.compile(r"\s*\([^)]*\)"),        # gemini-3-flash (thinking-minimal)（连同前导空格）
    re.compile(r"-\d{8}"),              # 20251101 日期快照
    re.compile(r"-\d{2}-\d{2}-\d{2}"),  # 26-02-10 日期快照（三位，须先于 MM-DD 规则）
    re.compile(r"-\d{2}-\d{2}$"),       # 12-10 日期快照（两位 MM-DD，锚定结尾避免误伤版本号）
    re.compile(r"-\d{4}"),              # 0110 / 2507 月日快照
    re.compile(r"-\d{1,3}k"),           # -32k 上下文档位
    re.compile(r"-latest"),
    re.compile(r"-(?:high|xhigh|max|medium|lite|low|thinking|reasoning|preview|beta\d*|exp)(?=-|$)"),
]


def normalize_family(name: str) -> str:
    """把 LMArena 模型名归并到「家族主干」：循环剥离档位/状态/日期快照后缀。"""
    key = str(name).lower().strip()
    for _ in range(6):
        new = key
        for rx in STRIP_RE:
            new = rx.sub("", new)
        new = re.sub(r"-+", "-", new).strip("- ").strip()
        if new == key:
            break
        key = new
    return key


def pretty_name(key: str) -> str:
    """家族主干 -> 展示名：claude-opus-5 -> Claude Opus 5；claude-opus-4-6 -> Claude Opus 4.6。"""
    # 把 "4-6" 这类版本号断句合并为 "4.6"，但避免误伤 "4-31b" 等参数量命名
    text = re.sub(r"(\d+)-(\d+)(?!\w)", r"\1.\2", key)
    parts = [p for p in text.split("-") if p]
    out = []
    for p in parts:
        if p in PROPER_NOUNS:
            out.append(PROPER_NOUNS[p])
        elif p[:1].isdigit():
            out.append(p)
        else:
            out.append(p[:1].upper() + p[1:])
    return " ".join(out)


def merge_families(rows: list[dict]) -> dict[str, dict]:
    """家族归并：同一家族保留 Elo 最高的条目（附带 org/license/展示名）。"""
    fam: dict[str, dict] = {}
    for r in rows:
        key = normalize_family(r["name"])
        cur = fam.get(key)
        if cur is None or r["rating"] > cur["rating"]:
            fam[key] = {
                "name": key,
                "display": pretty_name(key),
                "org": r["org"],
                "license": r["license"],
                "rating": r["rating"],
            }
    return fam


def match_model(src_name: str, mid: str) -> bool:
    """判断数据源模型名是否对应 models.json 中的 id。"""
    n = src_name.lower()
    for alias in MODEL_ALIASES.get(mid, []):
        if alias.lower() in n:
            return True
    return False


def apply_updates(models, elo_map, families, scores, prices) -> tuple[list[str], set[str]]:
    changed: list[str] = []
    covered: set[str] = set()  # 已被现有模型消费的家族键（避免重复新增）
    n_elo = n_sc = n_price = 0
    for m in models:
        mid = m["id"]
        # Elo
        hit = False
        for src_name, v in elo_map.items():
            if match_model(src_name, mid):
                new_v = round(v)
                old = m.get("elo")
                if old != new_v:
                    m["elo"] = new_v
                    changed.append(f"{mid}.elo {old}→{new_v}")
                m.get("est", {}).pop("elo", None)
                hit = True
                covered.add(normalize_family(src_name))
                break
        if not hit:
            # 未走别名时，尝试家族键直接匹配（全量同步新增的条目 id 即家族主干）
            fam = families.get(mid)
            if fam:
                new_v = round(fam["rating"])
                old = m.get("elo")
                if old != new_v:
                    m["elo"] = new_v
                    changed.append(f"{mid}.elo {old}→{new_v}")
                m.get("est", {}).pop("elo", None)
                hit = True
                covered.add(mid)
        if hit:
            n_elo += 1

        # SuperCLUE 智能指数 + 六维子分 + 综合能力分
        hit_sc = False
        for src_name, e in scores.items():
            if match_model(src_name, mid):
                total = e.get("total")
                if total is not None:
                    new_v = round(total, 2)
                    old = m.get("superclue")
                    if old != new_v:
                        m["superclue"] = new_v
                        changed.append(f"{mid}.superclue {old}→{new_v}")
                    m.get("est", {}).pop("superclue", None)
                dims = e.get("dims") or {}
                if dims:
                    m["dims"] = {k: round(v, 2) for k, v in dims.items()}
                    m.get("est", {}).pop("dims", None)
                    vals = [v for v in dims.values() if v is not None]
                    if vals:
                        b = round(sum(vals) / len(vals), 2)
                        old = m.get("bench")
                        if old != b:
                            m["bench"] = b
                            changed.append(f"{mid}.bench {old}→{b}")
                        m.get("est", {}).pop("bench", None)
                hit_sc = True
                break
        if hit_sc:
            n_sc += 1

        # SuperCLUE 价格
        hit_pr = False
        for src_name, e in prices.items():
            if match_model(src_name, mid):
                for field in ("priceIn", "priceOut"):
                    if field in e:
                        new_v = round(e[field], 2)
                        old = m.get(field)
                        if old != new_v:
                            m[field] = new_v
                            changed.append(f"{mid}.{field} {old}→{new_v}")
                        m.get("est", {}).pop(field, None)
                hit_pr = True
                break
        if hit_pr:
            n_price += 1

        # 清理空 est
        if m.get("est") == {}:
            m.pop("est", None)

    log(f"  ✓ 匹配更新：Elo {n_elo} 个模型，中文能力 {n_sc} 个模型，价格 {n_price} 个模型")
    if changed:
        log("  变更明细：")
        for c in changed:
            log(f"    - {c}")
    else:
        log("  数据无变化")
    return changed, covered


def sync_new_models(models: list[dict], families: dict[str, dict], covered: set[str]) -> int:
    """把 LMArena 上有、但 models.json 未收录的模型家族追加为条目。返回新增数。"""
    existing = {m["id"] for m in models}
    added = 0
    for key, fam in families.items():
        if key in covered or key in existing:
            continue
        org = fam["org"]
        provider = ORG_NAMES.get(org) or (org.title() if org else "其他")
        elo = round(fam["rating"])
        models.append({
            "id": key,
            "name": fam["display"],
            "provider": provider,
            "region": "china" if org in CN_ORGS else "overseas",
            "open": bool(fam["license"] and fam["license"].lower() != "proprietary"),
            "elo": elo,
            "desc": f"来自 {provider} 的模型，竞技场 Elo {elo} 分。数据每周自动更新。",
        })
        added += 1
    return added


def main() -> int:
    ap = argparse.ArgumentParser(description="MaaS Rank 数据自动更新")
    ap.add_argument("--data", default=str(DATA_FILE), help="models.json 路径")
    ap.add_argument("--dry-run", action="store_true", help="只打印变更不写文件")
    args = ap.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        log(f"✗ 找不到数据文件：{data_path}")
        return 1
    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)
    meta = data["meta"]

    today = date.today().isoformat()
    diff = {"updated": today, "sources": {}, "changes": []}

    # 1) LMArena
    elo_map, lm_rows = fetch_lmarena()
    diff["sources"]["lmarena"] = {"ok": bool(elo_map), "count": len(elo_map)}
    families = merge_families(lm_rows)

    # 2/3) SuperCLUE
    sc_scores, sc_prices, used_date = fetch_superclue()
    diff["sources"]["superclue"] = {
        "ok": bool(sc_scores),
        "count": len(sc_scores),
        "date": used_date,
    }

    # 合并（注意：apply_updates 会原地修改 models，因此先保留更新前快照用于历史归档）
    prev_snapshot = {
        "date": meta.get("updated", ""),
        "models": [
            {k: m.get(k) for k in HISTORY_FIELDS}
            for m in data["models"]
        ],
    }
    changes, covered = apply_updates(data["models"], elo_map, families, sc_scores, sc_prices)
    n_added = sync_new_models(data["models"], families, covered)
    if n_added:
        changes.append(f"+新增 {n_added} 个模型（家族归并后 LMArena 全量同步）")
        log(f"  ✓ 新增 {n_added} 个模型，当前共 {len(data['models'])} 个")
    # 按 Elo 降序重排（前端自行排序不受影响，仅便于人工审阅与快照对比）
    data["models"].sort(
        key=lambda m: m.get("elo") if m.get("elo") is not None else -1, reverse=True)
    diff["changes"] = changes

    # meta 更新（仅当确有数据变化时，避免空跑产生无意义提交）
    if changes:
        meta["updated"] = today
        meta["lastAutoUpdate"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sources_note = []
        if elo_map:
            sources_note.append("Elo 来自 LMArena（HuggingFace 数据集，自动抓取）")
        if sc_scores:
            sources_note.append(f"中文能力/价格来自 SuperCLUE（{used_date} 测评，自动抓取）")
        if sources_note:
            meta["note"] = "榜单数据自动更新：" + "；".join(sources_note) + "；未抓取到的字段保留上次快照。"

    if args.dry_run:
        log(f"（dry-run）共 {len(changes)} 处变更，未写文件。")
        return 0

    tmp = data_path.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(data_path)
    with open(DIFF_FILE, "w", encoding="utf-8") as f:
        json.dump(diff, f, ensure_ascii=False, indent=2)
    log(f"✓ 已写入 {data_path}")
    log(f"✓ 已写入变更摘要 {DIFF_FILE}")

    # 上周快照（用于首页「本周榜单变化」摘要卡片做对比）
    prev_path = ROOT / "data" / "prev.json"
    prev_tmp = prev_path.with_suffix(".json.tmp")
    with open(prev_tmp, "w", encoding="utf-8") as f:
        json.dump({"date": prev_snapshot["date"], "models": prev_snapshot["models"]},
                  f, ensure_ascii=False, indent=2)
    prev_tmp.replace(prev_path)
    log(f"✓ 已保存上周快照 {prev_path}（{len(prev_snapshot['models'])} 个模型）")

    # 历史快照归档：仅当数据确有变化时，把「更新前」的榜单写入 history.json，
    # 供榜单变化周报页对比展示。同一天重复更新则覆盖最后一条，避免堆积重复快照。
    if changes:
        history: dict = {"meta": {"count": 0, "generated": ""}, "snapshots": []}
        if HISTORY_FILE.exists():
            try:
                with open(HISTORY_FILE, encoding="utf-8") as f:
                    history = json.load(f)
            except Exception:  # noqa: BLE001
                log("  ⚠ history.json 读取失败，按空历史重建")
        snaps = history.setdefault("snapshots", [])
        if snaps and snaps[-1].get("date") == prev_snapshot.get("date"):
            snaps[-1] = prev_snapshot
        else:
            snaps.append(prev_snapshot)
        history["meta"] = {"count": len(snaps), "generated": today}
        htmp = HISTORY_FILE.with_suffix(".json.tmp")
        with open(htmp, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        htmp.replace(HISTORY_FILE)
        log(f"✓ 已归档历史快照 {len(snaps)} 期 → {HISTORY_FILE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
